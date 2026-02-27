"""
FitTracker — Flask backend
Stockage : JSON files (data/)
Charts    : Plotly (JSON renvoyé au front, rendu via plotly.js CDN)
"""

import json
import os
import uuid
from datetime import datetime, date
from pathlib import Path
from collections import defaultdict

from flask import Flask, jsonify, request, send_from_directory, abort
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── SETUP ─────────────────────────────────────────────────────────────────

import os as _os
_BASE = _os.path.dirname(_os.path.abspath(__file__))
app = Flask(__name__,
            static_folder=_os.path.join(_BASE, "static"),
            template_folder=_os.path.join(_BASE, "templates"))
DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)

EXERCISES_FILE = DATA_DIR / "exercises.json"
PROGRAMS_FILE  = DATA_DIR / "programs.json"
LOGS_FILE      = DATA_DIR / "logs.json"
EXCEL_FILE     = DATA_DIR / "fittracker_export.xlsx"

# ── PLOTLY DARK THEME ──────────────────────────────────────────────────────

LAYOUT_BASE = dict(
    paper_bgcolor="#111118",
    plot_bgcolor="#111118",
    font=dict(family="Inter, sans-serif", color="#8888aa", size=12),
    margin=dict(t=50, b=40, l=50, r=30),
    hovermode="x unified",
    hoverlabel=dict(bgcolor="#1e1e2a", font_color="#f0f0f8", bordercolor="#2a2a38"),
    legend=dict(bgcolor="#16161f", bordercolor="#2a2a38", borderwidth=1,
                font=dict(color="#c0c0d0", size=11)),
    colorway=["#e94560","#00d4ff","#ffd740","#00e676","#c792ea","#ff6b6b","#82aaff"],
)

AXIS_STYLE = dict(gridcolor="#2a2a38", linecolor="#2a2a38",
                  tickfont=dict(color="#8888aa"), zerolinecolor="#2a2a38")

def apply_theme(fig):
    fig.update_layout(**LAYOUT_BASE)
    fig.update_xaxes(**AXIS_STYLE)
    fig.update_yaxes(**AXIS_STYLE)
    return fig

def fig_json(fig):
    apply_theme(fig)
    return json.loads(fig.to_json())

# ── JSON HELPERS ───────────────────────────────────────────────────────────

def read_json(path: Path, default=None):
    if default is None:
        default = []
    try:
        if path.exists():
            return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        pass
    return default

def write_json(path: Path, data):
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def new_id():
    return str(uuid.uuid4())[:8]

def get_week_key(date_str):
    d = datetime.strptime(str(date_str)[:10], "%Y-%m-%d")
    iso = d.isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}"

# ── SEED DATA ──────────────────────────────────────────────────────────────

# Seed format: (name, muscle_group, category, description, tips, muscles_secondary, wger_id)
# wger_id = exercise ID on wger.de — used to fetch the animated GIF
# https://wger.de/api/v2/exerciseimage/?exercise={id}&format=json
EXERCISES_SEED = [
    # ── Pectoraux ────────────────────────────────────────────────────────────
    ("Développé couché",           "Pectoraux",  "free_weight",
     "Allongé sur un banc plat, saisissez la barre en prise large (légèrement au-delà des épaules). Descendez la barre jusqu'à effleurer votre poitrine, puis poussez verticalement en contractant les pectoraux. Gardez les pieds à plat, les omoplates rétractées et le dos légèrement cambré.",
     "Serrez la barre comme si vous vouliez la casser — ça active mieux les pecs. Expirer à la poussée.",
     "Triceps, Deltoïdes antérieurs", 192),

    ("Développé incliné haltères", "Pectoraux",  "free_weight",
     "Banc incliné à 30-45°. Haltères en prise neutre ou pronation, coudes à ~75° du torse. Descendez jusqu'à sentir l'étirement des pectoraux supérieurs, puis poussez en arc vers le haut. Excellente amplitude de mouvement.",
     "L'inclinaison cible davantage le faisceau claviculaire (haut des pecs). Ne pas monter le banc trop haut (>45°), ça sollicite trop les épaules.",
     "Triceps, Épaules antérieures", 314),

    ("Écarté haltères",            "Pectoraux",  "free_weight",
     "Allongé sur banc plat, haltères au-dessus de la poitrine, légère flexion des coudes fixe. Ouvrez les bras en arc de cercle jusqu'à ressentir l'étirement des pectoraux, puis ramenez en contractant fort en haut. Mouvement d'isolation pur.",
     "Ce n'est pas un exercice de force : utilisez un poids modéré, concentrez-vous sur la contraction. Gardez les coudes légèrement fléchis tout au long.",
     "Biceps (stabilisation)", 119),

    ("Pec Deck Machine",           "Pectoraux",  "machine",
     "Assis à la machine butterfly, coudes et avant-bras contre les coussinets. Ramenez les bras vers le centre en contractant les pectoraux, marquez une pause d'1 seconde, puis revenez lentement en contrôlant l'étirement.",
     "Parfait pour finir une séance pec avec une isolation totale. Le siège doit être réglé pour que les coudes soient à hauteur d'épaule.",
     "Épaules antérieures", 0),

    ("Chest Press Machine",        "Pectoraux",  "machine",
     "Dos bien appuyé contre le dossier, poignées à hauteur de poitrine. Poussez à fond en contractant les pecs, revenez lentement. La machine guide la trajectoire, idéale pour se concentrer sur la contraction sans se soucier de l'équilibre.",
     "Avantage sur le développé couché : pas besoin de pareur, tension constante tout au long du mouvement.",
     "Triceps, Épaules", 0),

    ("Cable Crossover",            "Pectoraux",  "machine",
     "Poulies hautes de chaque côté. En léger penchement vers l'avant, amenez les câbles vers le bas et vers l'intérieur en croisant légèrement les mains. Contraction maximale en bas, retour lent et contrôlé.",
     "La poulie maintient une tension constante sur les pectoraux, y compris en position contractée — avantage sur les haltères. Varier la hauteur des poulies cible différentes zones des pecs.",
     "Épaules antérieures, Biceps", 0),

    ("Dips",                       "Pectoraux",  "bodyweight",
     "Aux barres parallèles, corps légèrement penché en avant pour cibler les pecs (vertical = plus triceps). Descendez jusqu'à ce que les épaules soient sous le niveau des coudes, remontez en poussant. Amplitude complète.",
     "Penchez-vous vers l'avant et écartez légèrement les coudes pour maximiser le travail des pectoraux. Descendre sous 90° si la mobilité le permet.",
     "Triceps, Épaules antérieures", 73),

    ("Pompes",                     "Pectoraux",  "bodyweight",
     "Corps en planche rigide, mains légèrement plus larges que les épaules. Descendez la poitrine jusqu'au sol en gardant les coudes à ~45° du corps, poussez jusqu'à extension complète en contractant les pecs.",
     "Serrez les abdos et les fessiers pendant tout le mouvement pour maintenir le gainage. Large = plus de pecs, serré = plus de triceps.",
     "Triceps, Épaules, Abdominaux", 35),

    # ── Dos ──────────────────────────────────────────────────────────────────
    ("Tractions",                  "Dos",        "bodyweight",
     "Prise pronation (pull-up) ou supination (chin-up), mains légèrement plus larges que les épaules. Partez en suspension complète, tirez jusqu'au menton au-dessus de la barre en rétractant les omoplates. Descendez lentement.",
     "Imaginez que vous tirez des coudes vers les hanches plutôt que de fléchir les bras — ça recrute mieux le grand dorsal. Prise supination = plus de biceps.",
     "Biceps, Rhomboïdes, Grand rond", 31),

    ("Rowing haltère",             "Dos",        "free_weight",
     "Un genou et une main sur un banc. Dos plat, haltère dans la main libre, bras tendu. Tirez l'haltère vers la hanche en rétractant l'omoplate, coude le long du corps. Descendez lentement en gardant le contrôle.",
     "Penser à 'mettre la main dans la poche arrière' pour mieux activer le grand dorsal. Éviter de tourner le tronc.",
     "Biceps, Rhomboïdes, Trapèzes", 197),

    ("Soulevé de terre",           "Dos",        "free_weight",
     "Pieds dans la largeur des hanches, barre au-dessus du milieu du pied. Saisissez la barre en prise alternée ou double pronation. Dos plat, hanches plus hautes que les genoux, épaules au-dessus de la barre. Poussez le sol et tirez la barre en gardant le dos neutre.",
     "Le roi des exercices. Gardez la barre le long des tibias/cuisses tout au long du mouvement. Inspirez profondément et bloquez avant de tirer (valsalva).",
     "Quadriceps, Fessiers, Trapèzes, Ischio-jambiers", 29),

    ("Rowing barre",               "Dos",        "free_weight",
     "Penché à ~45°, barre en prise pronation. Tirez la barre vers le nombril/bas du sternum en rétractant les omoplates. Coudes vers l'arrière, pas vers l'extérieur. Descendez lentement.",
     "La position inclinée (45°) recrute plus le grand dorsal. Plus vertical (70°+) = plus de trapèzes supérieurs.",
     "Biceps, Trapèzes, Rhomboïdes", 63),

    ("Lat Pulldown",               "Dos",        "machine",
     "Assis, cuisses bloquées sous les genouillères. Prise pronation large, tirez la barre vers le haut de la poitrine en rétractant les omoplates et en sortant la poitrine vers la barre. Revenez lentement bras tendus.",
     "Inclinez légèrement le buste en arrière (10-15°) pour une meilleure amplitude. Tirez vers la clavicule, pas derrière la nuque (risque cervical).",
     "Biceps, Grand rond, Rhomboïdes", 122),

    ("Rowing assis machine",       "Dos",        "machine",
     "Poitrine contre le support, saisissez les poignées. Tirez vers vous en rétractant les omoplates et en serrant le dos en fin de mouvement, revenez lentement avec un bon étirement. Dos droit tout au long.",
     "Varier la prise : prise serrée neutre = plus de grand dorsal, prise large = plus de rhomboïdes/trapèzes.",
     "Biceps, Rhomboïdes, Trapèzes", 115),

    ("Face Pull",                  "Dos",        "machine",
     "Poulie haute avec corde. Tirez la corde vers votre visage, coudes écartés à hauteur d'épaules, en tournant les poignets vers l'extérieur à l'arrivée. Excellent pour la santé des épaules et les trapèzes moyens/inférieurs.",
     "Exercice indispensable pour équilibrer le travail de poussée et protéger les épaules. À faire à chaque séance pull.",
     "Deltoïdes postérieurs, Trapèzes, Rotateurs externes", 0),

    # ── Épaules ───────────────────────────────────────────────────────────────
    ("Développé militaire",        "Épaules",    "free_weight",
     "Debout ou assis, barre devant à hauteur de la clavicule. Poussez la barre verticalement au-dessus de la tête jusqu'à extension complète des bras. Descendez lentement. Core solide pour protéger le bas du dos.",
     "Rentrez le menton quand la barre passe devant le visage (pas la tête en arrière). Serrez les fessiers pour stabiliser le bassin et éviter de cambrer.",
     "Triceps, Trapèzes, Dentelé antérieur", 68),

    ("Élévations latérales",       "Épaules",    "free_weight",
     "Debout, haltères le long du corps. Levez les bras sur les côtés jusqu'à l'horizontale, légère rotation externe (le petit doigt légèrement plus haut). Descendez lentement. Coudes très légèrement fléchis.",
     "Le mouvement clé pour élargir les épaules. Poids léger, haute répétition. Éviter de balancer le tronc — si vous le faites, le poids est trop lourd.",
     "Trapèzes supérieurs (si trop lourd)", 0),

    ("Élévations frontales",       "Épaules",    "free_weight",
     "Haltères ou barre, levez les bras droit devant vous jusqu'à l'horizontale, pouce légèrement vers le haut. Descendez lentement. Isolement du faisceau antérieur (avant) des deltoïdes.",
     "Souvent déjà bien sollicité par les développés. Poids modéré, contrôle maximal. Alterner les bras possible.",
     "Pectoraux supérieurs", 0),

    ("Arnold Press",               "Épaules",    "free_weight",
     "Assis, haltères devant le visage, paumes vers soi. En montant, tournez progressivement les paumes vers l'avant et poussez au-dessus de la tête. Inversez le mouvement en descendant. Recrute les 3 faisceaux des deltoïdes.",
     "Inventé par Arnold Schwarzenegger. La rotation implique les deltoïdes antérieurs et moyens ainsi que les trapèzes. Mouvement complet mais complexe.",
     "Triceps, Trapèzes", 0),

    ("Shoulder Press Machine",     "Épaules",    "machine",
     "Assis, dos calé. Poignées à hauteur des épaules, poussez vers le haut jusqu'à extension sans verrouiller les coudes, revenez lentement. La machine guide la trajectoire.",
     "Idéal pour s'entraîner lourd en sécurité sans pareur. Régler le siège pour que les poignées soient à hauteur des oreilles.",
     "Triceps, Trapèzes", 0),

    # ── Biceps ────────────────────────────────────────────────────────────────
    ("Curl barre",                 "Biceps",     "free_weight",
     "Debout, barre en prise supination (paumes vers le haut). Fléchissez les coudes en remontant la barre vers les épaules, coudes fixes le long du corps. Descendez lentement avec résistance.",
     "Les coudes NE bougent PAS. Imaginez qu'ils sont vissés à vos hanches. En haut, contractez fort 1 seconde.",
     "Brachial, Brachio-radial", 11),

    ("Curl haltères",              "Biceps",     "free_weight",
     "Debout ou assis, haltères en prise supination. Fléchissez un bras à la fois ou simultanément, en tournant le poignet vers l'extérieur en montant (supination). Descendez lentement.",
     "La supination (rotation du poignet en montant) maximise l'activation du biceps. Plus grande amplitude de mouvement que la barre.",
     "Brachial, Brachio-radial", 42),

    ("Curl marteau",               "Biceps",     "free_weight",
     "Haltères en prise neutre (poignets dans l'axe des avant-bras, comme si vous teniez un marteau). Fléchissez les coudes en montant, coudes fixes. Travaille principalement le muscle brachial et le brachio-radial.",
     "Plus facile pour les poignets que le curl classique. Excellent pour épaissir le bras et renforcer les avant-bras.",
     "Brachio-radial, Avant-bras", 0),

    ("Curl poulie basse",          "Biceps",     "machine",
     "Face à la poulie basse, barre ou corde. Tirez vers le haut en fléchissant les coudes, coudes fixes. La poulie maintient une tension constante sur les biceps, y compris en position basse (contrairement aux haltères).",
     "Excellent exercice de finition. La tension constante favorise le pump et l'hypertrophie.",
     "Brachial", 0),

    # ── Triceps ───────────────────────────────────────────────────────────────
    ("Skull Crusher",              "Triceps",    "free_weight",
     "Allongé sur banc plat, barre ou haltères au-dessus de la tête, bras tendus. Fléchissez uniquement les coudes en descendant la barre vers le front (ou derrière la tête), puis remontez. Coudes pointés vers le plafond.",
     "Aussi appelé 'French press couché'. Le nom dit tout sur la vigilance requise. Descente lente et contrôlée. Excellent pour le chef long du triceps.",
     "Chef long du triceps (etirement maximal)", 26),

    ("Extension au-dessus tête",   "Triceps",    "free_weight",
     "Debout ou assis, haltère ou barre au-dessus de la tête, bras tendus. Fléchissez les coudes pour descendre le poids derrière la tête, puis remontez. Maximise l'étirement du chef long du triceps.",
     "Le chef long du triceps est le seul qui s'étire complètement avec le bras au-dessus. Mouvement clé pour l'hypertrophie du triceps.",
     "Chef long du triceps", 0),

    ("Dips triceps",               "Triceps",    "bodyweight",
     "Aux barres parallèles ou entre deux bancs. Corps vertical (pas penché en avant). Descendez jusqu'à 90° des coudes, remontez. Coudes près du corps pour cibler les triceps.",
     "Corps vertical = triceps. Corps penché = pectoraux. Excellent exercice de masse pour les triceps.",
     "Pectoraux (si penché), Épaules", 73),

    ("Extension poulie haute",     "Triceps",    "machine",
     "Face à la poulie haute, corde ou barre droite. Coudes fixes près du corps, poussez vers le bas jusqu'à extension complète, tournez légèrement les poignets vers l'extérieur en bas (avec corde). Remontez lentement.",
     "Coudes collés aux flancs — ils NE bougent PAS. C'est l'erreur la plus commune. Variante corde = meilleure contraction en bas.",
     "Chef latéral et médial du triceps", 86),

    # ── Jambes ────────────────────────────────────────────────────────────────
    ("Squat",                      "Jambes",     "free_weight",
     "Barre sur les trapèzes (ou haut du dos), pieds dans la largeur des épaules, pointes légèrement ouvertes. Descendez en poussant les genoux dans l'axe des pieds, dos droit, jusqu'à ce que les cuisses soient parallèles ou en dessous. Remontez en poussant fort dans le sol.",
     "Le roi des exercices de jambes. Les genoux suivent la direction des orteils. Regardez légèrement vers le bas-devant. Respirez profondément avant de descendre.",
     "Fessiers, Ischio-jambiers, Mollets, Dos", 29),

    ("Leg Press",                  "Jambes",     "machine",
     "Dos calé dans le siège, pieds sur la plateforme à largeur d'épaules. Descendez lentement jusqu'à 90° (ou plus si la mobilité le permet), poussez la plateforme sans verrouiller les genoux en haut.",
     "Pieds hauts sur la plateforme = plus de fessiers/ischios. Pieds bas = plus de quadriceps. Excellente alternative au squat pour travailler lourd sans charge axiale.",
     "Fessiers, Ischio-jambiers", 0),

    ("Leg Extension",              "Jambes",     "machine",
     "Assis, dos calé, chevilles sous les coussinets. Tendez les jambes jusqu'à l'horizontale en contractant fort les quadriceps, puis descendez lentement. Isolation pure des quadriceps.",
     "Exercice d'isolation. Pointer les orteils légèrement vers l'extérieur peut aider à mieux sentir le vaste interne.",
     "Quadriceps uniquement", 105),

    ("Leg Curl couché",            "Jambes",     "machine",
     "Allongé face vers le bas, chevilles sous les coussinets. Fléchissez les genoux en ramenant les talons vers les fessiers, contractez fort les ischio-jambiers, descendez lentement. Hanche bien appuyée contre le coussin.",
     "Étirement complet en bas = plus d'hypertrophie. Tourner les pointes des pieds légèrement vers l'intérieur active mieux le biceps fémoral.",
     "Ischio-jambiers, Mollets (secondaire)", 116),

    ("Romanian Deadlift",          "Jambes",     "free_weight",
     "Debout, barre ou haltères devant les cuisses. Poussez les hanches vers l'arrière (hip hinge) en descendant la barre le long des jambes, dos parfaitement plat, jusqu'à sentir l'étirement des ischios (généralement mi-tibia). Remontez en poussant les hanches vers l'avant.",
     "Ce n'est PAS un squat : les genoux restent presque droits. Le mouvement vient des hanches. Sensation d'étirement intense dans l'arrière des cuisses = bonne exécution.",
     "Fessiers, Bas du dos", 0),

    ("Fentes",                     "Jambes",     "free_weight",
     "Debout, haltères en main. Faites un grand pas en avant, descendez jusqu'à ce que le genou arrière effleure le sol. Le genou avant reste au-dessus du pied. Remontez et alternez les jambes.",
     "Excellent pour corriger les déséquilibres gauche/droite. Les fentes marchées (en déplacement) sont plus fonctionnelles et augmentent la difficulté.",
     "Fessiers, Ischio-jambiers", 0),

    ("Hip Thrust",                 "Fessiers",   "free_weight",
     "Dos appuyé sur un banc (hauteur des omoplates), barre sur les hanches (avec pad). Pieds à plat, dans la largeur des hanches. Poussez les hanches vers le haut jusqu'à former une ligne droite genou-hanche-épaule. Contractez fort les fessiers en haut.",
     "L'exercice le plus efficace pour les fessiers. Contractez FORT en haut et maintenez 1 seconde. Inclinez le menton vers la poitrine pour protéger le cou.",
     "Ischio-jambiers, Quadriceps", 0),

    ("Mollets debout machine",     "Mollets",    "machine",
     "Debout sur la plateforme (avant du pied sur le rebord), épaules sous les coussinets. Montez sur la pointe des pieds le plus haut possible, contractez 1s, descendez jusqu'à l'étirement maximum.",
     "L'amplitude complète est clé : étirement total en bas (talon très bas) et contraction maximale en haut. Les mollets répondent bien aux séries longues (15-20 reps).",
     "Soléaire (si genou fléchi)", 0),

    ("Mollets assis machine",      "Mollets",    "machine",
     "Assis, genoux fléchis à 90° sous les coussinets, avant du pied sur la plateforme. Montez sur la pointe des pieds, contractez, puis descendez en étirement complet. Le genou fléchi isole le soléaire.",
     "Genou fléchi = soléaire dominant. Genou tendu = gastrocnémien dominant. Les deux sont nécessaires pour des mollets complets.",
     "Gastrocnémien (secondaire)", 0),

    # ── Abdominaux ────────────────────────────────────────────────────────────
    ("Crunch",                     "Abdominaux", "bodyweight",
     "Allongé sur le dos, genoux fléchis, mains derrière la tête (sans tirer sur le cou). Enroulez le haut du tronc vers les genoux en expirant, en contractant les abdos. Descendez lentement sans poser complètement la tête.",
     "Mouvement court et contrôlé. Le bas du dos reste au sol. Ce n'est pas un sit-up complet. Expirez fort en montant.",
     "Abdominaux droits superficiels", 0),

    ("Planche",                     "Abdominaux", "bodyweight",
     "En appui sur les avant-bras et les orteils, corps en ligne droite de la tête aux talons. Contractez abdominaux, fessiers et quadriceps simultanément. Maintenez la position.",
     "Regardez 30cm devant vous. Les hanches ne doivent ni monter ni descendre. Respiration normale. Progressez en durée : 20s → 30s → 60s → 2min.",
     "Dorsaux, Fessiers, Quadriceps", 0),

    ("Relevé de jambes",           "Abdominaux", "bodyweight",
     "Suspendu à une barre ou allongé. Levez les jambes tendues (ou fléchies pour commencer) jusqu'à l'horizontale ou plus. Descendez lentement sans balancement. Sollicite fortement le bas des abdominaux et le psoas.",
     "En suspension = plus difficile et plus efficace. Éviter le balancement en contrôlant chaque répétition. Jambes tendues = plus intense.",
     "Psoas, Hip-flexors", 0),

    ("Russian Twist",              "Abdominaux", "bodyweight",
     "Assis, buste incliné à 45°, pieds au sol ou levés. Tournez le buste de gauche à droite en amenant les mains (ou un poids) d'un côté à l'autre. Travaille les obliques et la rotation du tronc.",
     "Plus les pieds sont levés, plus c'est difficile. Ajouter un poids ou un ballon médicinal pour plus d'intensité.",
     "Obliques, Bas du dos", 0),

    ("Crunch poulie haute",        "Abdominaux", "machine",
     "À genoux face à la poulie haute avec corde. Saisissez la corde de chaque côté de la tête. Enroulez le buste vers le bas en contractant les abdos, coudes qui vont vers les genoux. Ne pas tirer avec les bras.",
     "La résistance de la poulie permet de surcharger progressivement les abdominaux comme n'importe quel autre muscle. Excellente alternative aux crunchs classiques.",
     "Abdominaux droits, Obliques", 0),

    # ── Calisthenics ─────────────────────────────────────────────────────────
    ("Muscle Up",                  "Full Body",  "calisthenics",
     "Traction explosive jusqu'au-delà de la barre, puis transition en dips. Mouvement en deux phases : phase pull (tirage) et phase push (poussée au-dessus). Requiert une forte traction et une bonne coordination.",
     "La transition est la partie la plus difficile. Travailler les 'negative muscle-up' pour apprendre. La prise en faux (false grip) aide pour la version aux anneaux.",
     "Dos, Pectoraux, Triceps, Core", 0),

    ("Front Lever",                "Dos",        "calisthenics",
     "Suspendu à une barre, corps horizontal face vers le haut, bras tendus. Le corps forme une planche rigide. Progression : genoux fléchis → une jambe tendue → deux jambes tendues.",
     "Exercice statique de force extraordinaire. Contracter le grand dorsal comme si vous vouliez casser la barre. Progressions : tuck → advanced tuck → one leg → straddle → full.",
     "Biceps, Core, Grand dorsal", 0),

    ("Back Lever",                 "Dos",        "calisthenics",
     "Suspendu à une barre, corps horizontal face vers le bas, bras tendus derrière. Tenu statiquement ou utilisé comme passage dans des mouvements gymniques.",
     "Moins difficile que le front lever mais requiert une bonne mobilité d'épaule. Progressions similaires : tuck → advanced tuck → one leg → full.",
     "Pectoraux, Biceps, Core", 0),

    ("Planche (figure)",           "Épaules",    "calisthenics",
     "En appui sur les deux mains, corps horizontal face vers le bas, bras tendus. L'une des figures les plus difficiles de la calisthenics. Requiert une force extraordinaire des épaules et du core.",
     "Progression : planche lean → tuck planche → advanced tuck → straddle → full. Compter en années d'entraînement. La pseudo-planche push-up est un bon travail préparatoire.",
     "Core, Triceps, Grand dorsal", 0),

    ("Human Flag",                 "Full Body",  "calisthenics",
     "Corps horizontal, perpendiculaire à un poteau vertical, maintenu par la force des bras et du côté du corps. Une main pousse, l'autre tire.",
     "Figure spectaculaire nécessitant une force latérale exceptionnelle. Progressions : vertical flag → diagonal → tucked → full. Travailler les lateral press au sol pour progresser.",
     "Obliques, Épaules, Grand dorsal, Core", 0),

    ("L-Sit",                      "Abdominaux", "calisthenics",
     "En appui sur deux barres ou le sol, corps soulevé, jambes tendues horizontalement formant un L. Position statique tenue.",
     "Progressions : L-sit genoux fléchis → une jambe → full L-sit → V-sit. Travailler la mobilité des hanches et des ischios en parallèle.",
     "Psoas, Quadriceps, Triceps", 0),

    ("Dragon Flag",                "Abdominaux", "calisthenics",
     "Allongé sur un banc, tenant une barre derrière la tête. Le corps monte en appui sur les épaules, puis descend lentement en position horizontale, parfaitement rigide. Exercice de Bruce Lee.",
     "L'un des exercices abdominaux les plus difficiles. Descente excentrique lente = progrès rapides. Corps rigide comme une planche.",
     "Core complet, Psoas, Épaules", 0),

    ("Pseudo Planche Push Up",     "Pectoraux",  "calisthenics",
     "Position pompes mais avec les doigts pointés vers les pieds et le poids sur les paumes. Corps légèrement incliné vers l'avant. Mouvement de pompes dans cette position.",
     "Excellent travail préparatoire pour la vraie planche. Plus l'inclinaison vers l'avant est grande, plus c'est difficile et proche des exigences de la planche.",
     "Épaules, Triceps, Core", 0),

    ("Archer Pull Up",             "Dos",        "calisthenics",
     "Tractions avec un bras qui tire et l'autre qui aide avec un bras tendu sur le côté. Le bras tendu est en supination, bras écartés. Transition vers la traction un bras.",
     "Excellent intermédiaire vers le one arm pull-up. Plus le bras assistant est tendu, plus c'est difficile.",
     "Grand dorsal, Biceps, Rhomboïdes", 0),

    ("Handstand Push Up",          "Épaules",    "calisthenics",
     "En équilibre sur les mains (dos au mur pour commencer), descendez la tête vers le sol en fléchissant les coudes, puis remontez. Corps le plus vertical possible.",
     "Progressions : wall HSPU → chest to wall → freestanding. Le mouvement le plus efficace pour les épaules en calisthenics.",
     "Triceps, Trapèzes", 0),

    ("One Arm Pull Up",            "Dos",        "calisthenics",
     "Traction stricte avec un seul bras. L'autre bras reste le long du corps ou dans le dos. Représente plusieurs années d'entraînement pour la plupart.",
     "Progressions : assisted OAP avec bande → archer pull-up → negative OAP → full OAP. La force du core est aussi cruciale.",
     "Grand dorsal, Biceps, Core", 0),

    ("Explosive Pull Up",          "Dos",        "calisthenics",
     "Traction avec une phase concentrique explosive pour amener le buste bien au-dessus de la barre, voire lâcher les mains au sommet. Travaille la puissance.",
     "Base du muscle-up. Progressions : traction explosive poitrine à la barre → lâcher et rattraper → muscle-up.",
     "Grand dorsal, Biceps, Trapèzes", 0),

    ("Tractions lestées",          "Dos",        "calisthenics",
     "Tractions classiques avec un poids supplémentaire (ceinture lestée, gilet, haltère entre les pieds). Même technique que les tractions normales, pleine amplitude.",
     "Une fois que vous faites 15+ tractions strictes, ajouter du poids pour continuer à progresser en force.",
     "Biceps, Rhomboïdes, Grand rond", 0),

    ("Dips lestés",                "Triceps",    "calisthenics",
     "Dips aux barres parallèles avec poids supplémentaire. Corps légèrement penché pour travailler les pectoraux, ou vertical pour les triceps.",
     "Même logique que les tractions lestées. Ajouter du poids progressivement une fois que les dips classiques deviennent faciles.",
     "Pectoraux, Épaules antérieures", 0),
]

PROGRAMS_SEED = [
    {
        "name": "Full Body 3j/semaine",
        "description": "Programme complet 3 séances par semaine, idéal débutant/intermédiaire",
        "sessions_per_week": 3,
        "sessions": [
            ("Séance A — Push", "Lundi",    [("Développé couché",4,"8-10"),("Développé militaire",3,"10-12"),("Dips",3,"Max"),("Élévations latérales",3,"12-15"),("Extension poulie haute",3,"12-15")]),
            ("Séance B — Pull", "Mercredi", [("Tractions",4,"Max"),("Rowing haltère",3,"10-12"),("Lat Pulldown",3,"10-12"),("Curl barre",3,"10-12"),("Face Pull",3,"15-20")]),
            ("Séance C — Legs", "Vendredi", [("Squat",4,"8-10"),("Leg Press",3,"12-15"),("Romanian Deadlift",3,"10-12"),("Leg Curl couché",3,"12-15"),("Mollets debout machine",4,"15-20")]),
        ],
    },
    {
        "name": "PPL 6j/semaine",
        "description": "Push Pull Legs 6 jours, niveau intermédiaire / avancé",
        "sessions_per_week": 6,
        "sessions": [
            ("Push A", "Lundi",    [("Développé couché",4,"6-8"),("Développé incliné haltères",3,"10-12"),("Écarté haltères",3,"12-15"),("Développé militaire",3,"8-10"),("Élévations latérales",4,"12-15"),("Skull Crusher",3,"10-12")]),
            ("Pull A", "Mardi",    [("Soulevé de terre",4,"5-6"),("Tractions",4,"Max"),("Rowing barre",3,"8-10"),("Lat Pulldown",3,"10-12"),("Curl barre",3,"10-12"),("Curl marteau",3,"12-15")]),
            ("Legs A", "Mercredi", [("Squat",5,"5-8"),("Leg Press",3,"12-15"),("Romanian Deadlift",3,"10"),("Leg Curl couché",3,"12"),("Mollets debout machine",5,"15")]),
            ("Push B", "Jeudi",    [("Chest Press Machine",4,"8-10"),("Dips",4,"10-12"),("Cable Crossover",3,"15"),("Arnold Press",3,"10-12"),("Élévations frontales",3,"12-15"),("Extension au-dessus tête",3,"12")]),
            ("Pull B", "Vendredi", [("Rowing haltère",4,"10-12"),("Face Pull",4,"15-20"),("Rowing assis machine",3,"12"),("Curl poulie basse",3,"12-15"),("Curl haltères",3,"12")]),
            ("Legs B", "Samedi",   [("Fentes",4,"12"),("Hip Thrust",4,"12-15"),("Leg Extension",3,"15"),("Leg Curl couché",3,"15"),("Mollets assis machine",5,"20")]),
        ],
    },
    {
        "name": "Calisthenics Débutant",
        "description": "Street workout — construire les bases pour les figures avancées",
        "sessions_per_week": 3,
        "sessions": [
            ("Fondamentaux Push", "Lundi",    [("Pompes",4,"Max"),("Dips",4,"Max"),("Pseudo Planche Push Up",3,"8-10"),("Planche",3,"30-60s")]),
            ("Fondamentaux Pull", "Mercredi", [("Tractions",5,"Max"),("Archer Pull Up",3,"6-8"),("L-Sit",4,"20-30s"),("Relevé de jambes",4,"15")]),
            ("Figures & Force",   "Vendredi", [("Muscle Up",4,"Max"),("Dragon Flag",3,"8-10"),("Squat",4,"20"),("Fentes",3,"15")]),
        ],
    },
    {
        "name": "Upper / Lower 4j/semaine",
        "description": "Haut / Bas du corps, 4 séances, force + hypertrophie",
        "sessions_per_week": 4,
        "sessions": [
            ("Upper A — Force",        "Lundi",    [("Développé couché",4,"5-6"),("Tractions",4,"6-8"),("Développé militaire",3,"6-8"),("Rowing barre",3,"6-8"),("Dips lestés",3,"8-10")]),
            ("Lower A — Force",        "Mardi",    [("Squat",5,"5"),("Romanian Deadlift",3,"8"),("Leg Press",3,"10"),("Leg Curl couché",3,"10"),("Mollets debout machine",4,"12")]),
            ("Upper B — Hypertrophie", "Jeudi",    [("Développé incliné haltères",4,"10-12"),("Lat Pulldown",4,"10-12"),("Écarté haltères",3,"12-15"),("Rowing haltère",3,"12"),("Curl barre",3,"12"),("Extension poulie haute",3,"12")]),
            ("Lower B — Hypertrophie", "Vendredi", [("Leg Press",4,"12-15"),("Fentes",3,"12"),("Leg Extension",3,"15"),("Hip Thrust",3,"12-15"),("Mollets assis machine",4,"20")]),
        ],
    },
]


def init_data():
    if not EXERCISES_FILE.exists():
        exercises = [
            {"id": new_id(), "name": n, "muscle_group": mg, "category": cat,
             "description": desc, "tips": tips, "muscles_secondary": sec,
             "wger_id": wger_id, "is_custom": False}
            for n, mg, cat, desc, tips, sec, wger_id in EXERCISES_SEED
        ]
        write_json(EXERCISES_FILE, exercises)

    if not PROGRAMS_FILE.exists():
        exercises = read_json(EXERCISES_FILE, [])
        ex_by_name = {e["name"]: e["id"] for e in exercises}
        programs = []
        for p in PROGRAMS_SEED:
            prog = {
                "id": new_id(), "name": p["name"], "description": p["description"],
                "sessions_per_week": p["sessions_per_week"], "is_custom": False,
                "created_at": str(date.today()), "sessions": [],
            }
            for sname, sday, sexs in p["sessions"]:
                session = {"id": new_id(), "name": sname, "day_of_week": sday, "exercises": []}
                for ename, sets, reps in sexs:
                    ex_id = ex_by_name.get(ename)
                    if ex_id:
                        session["exercises"].append({"id": new_id(), "exercise_id": ex_id,
                                                     "sets": sets, "target_reps": reps})
                prog["sessions"].append(session)
            programs.append(prog)
        write_json(PROGRAMS_FILE, programs)

    if not LOGS_FILE.exists():
        write_json(LOGS_FILE, [])


# ── STATIC ─────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    from flask import render_template
    return render_template("index.html")


# ── EXERCISES ─────────────────────────────────────────────────────────────

@app.route("/api/exercises", methods=["GET"])
def get_exercises():
    exs = read_json(EXERCISES_FILE, [])
    cat = request.args.get("category")
    if cat:
        exs = [e for e in exs if e.get("category") == cat]
    return jsonify(exs)

@app.route("/api/exercises", methods=["POST"])
def create_exercise():
    d = request.json
    exs = read_json(EXERCISES_FILE, [])
    ex = {"id": new_id(), "name": d["name"], "muscle_group": d.get("muscle_group",""),
          "category": d.get("category","free_weight"), "description": d.get("description",""),
          "is_custom": True}
    exs.append(ex)
    write_json(EXERCISES_FILE, exs)
    return jsonify(ex)

@app.route("/api/exercises/<ex_id>", methods=["DELETE"])
def delete_exercise(ex_id):
    exs = [e for e in read_json(EXERCISES_FILE, []) if e["id"] != ex_id]
    write_json(EXERCISES_FILE, exs)
    return jsonify({"ok": True})


# ── PROGRAMS ──────────────────────────────────────────────────────────────

def _enrich_programs(programs, exercises):
    ex_map = {e["id"]: e for e in exercises}
    for p in programs:
        for s in p.get("sessions", []):
            for se in s.get("exercises", []):
                ex = ex_map.get(se["exercise_id"], {})
                se["exercise_name"] = ex.get("name", "?")
                se["muscle_group"]  = ex.get("muscle_group", "")
                se["category"]      = ex.get("category", "")
    return programs

@app.route("/api/programs", methods=["GET"])
def get_programs():
    programs  = read_json(PROGRAMS_FILE, [])
    exercises = read_json(EXERCISES_FILE, [])
    return jsonify(_enrich_programs(programs, exercises))

@app.route("/api/programs", methods=["POST"])
def create_program():
    d = request.json
    programs = read_json(PROGRAMS_FILE, [])
    prog = {"id": new_id(), "name": d["name"], "description": d.get("description",""),
            "sessions_per_week": d.get("sessions_per_week", 3), "is_custom": True,
            "created_at": str(date.today()), "sessions": []}
    programs.append(prog)
    write_json(PROGRAMS_FILE, programs)
    return jsonify(prog)

@app.route("/api/programs/<pid>", methods=["DELETE"])
def delete_program(pid):
    programs = [p for p in read_json(PROGRAMS_FILE, []) if p["id"] != pid]
    write_json(PROGRAMS_FILE, programs)
    return jsonify({"ok": True})

@app.route("/api/programs/<pid>/sessions", methods=["POST"])
def add_session(pid):
    d = request.json
    programs = read_json(PROGRAMS_FILE, [])
    prog = next((p for p in programs if p["id"] == pid), None)
    if not prog: abort(404)
    session = {"id": new_id(), "name": d["name"], "day_of_week": d.get("day_of_week",""), "exercises": []}
    prog["sessions"].append(session)
    write_json(PROGRAMS_FILE, programs)
    return jsonify(session)

@app.route("/api/sessions/<sid>", methods=["DELETE"])
def delete_session(sid):
    programs = read_json(PROGRAMS_FILE, [])
    for p in programs:
        p["sessions"] = [s for s in p["sessions"] if s["id"] != sid]
    write_json(PROGRAMS_FILE, programs)
    return jsonify({"ok": True})

@app.route("/api/sessions/<sid>/exercises", methods=["POST"])
def add_session_exercise(sid):
    d = request.json
    programs = read_json(PROGRAMS_FILE, [])
    for p in programs:
        for s in p["sessions"]:
            if s["id"] == sid:
                se = {"id": new_id(), "exercise_id": d["exercise_id"],
                      "sets": d.get("sets",3), "target_reps": d.get("target_reps","8-12")}
                s["exercises"].append(se)
                write_json(PROGRAMS_FILE, programs)
                return jsonify(se)
    abort(404)

@app.route("/api/session_exercises/<se_id>", methods=["DELETE"])
def delete_session_exercise(se_id):
    programs = read_json(PROGRAMS_FILE, [])
    for p in programs:
        for s in p["sessions"]:
            s["exercises"] = [e for e in s["exercises"] if e["id"] != se_id]
    write_json(PROGRAMS_FILE, programs)
    return jsonify({"ok": True})


# ── LOGS ──────────────────────────────────────────────────────────────────

def _enrich_logs(logs):
    exercises = read_json(EXERCISES_FILE, [])
    programs  = read_json(PROGRAMS_FILE, [])
    ex_map    = {e["id"]: e for e in exercises}
    prog_map  = {p["id"]: p["name"] for p in programs}
    sess_map  = {s["id"]: s["name"] for p in programs for s in p["sessions"]}
    for log in logs:
        log["program_name"] = prog_map.get(log.get("program_id"), "")
        log["session_name"] = sess_map.get(log.get("session_id"), "Séance libre")
        for s in log.get("sets", []):
            ex = ex_map.get(s.get("exercise_id"), {})
            s["exercise_name"] = ex.get("name", "?")
            s["category"]      = ex.get("category", "")
    return logs

@app.route("/api/logs", methods=["GET"])
def get_logs():
    logs = read_json(LOGS_FILE, [])
    logs = sorted(logs, key=lambda l: l.get("log_date",""), reverse=True)
    limit = int(request.args.get("limit", 200))
    return jsonify(_enrich_logs(logs[:limit]))

@app.route("/api/logs", methods=["POST"])
def create_log():
    d = request.json
    logs = read_json(LOGS_FILE, [])
    log = {"id": new_id(), "session_id": d.get("session_id"), "program_id": d.get("program_id"),
           "log_date": d.get("log_date", str(date.today())), "notes": d.get("notes",""),
           "created_at": datetime.now().isoformat(), "sets": d.get("sets",[])}
    logs.append(log)
    write_json(LOGS_FILE, logs)
    _export_to_excel()
    return jsonify(log)

@app.route("/api/logs/<lid>", methods=["DELETE"])
def delete_log(lid):
    logs = [l for l in read_json(LOGS_FILE, []) if l["id"] != lid]
    write_json(LOGS_FILE, logs)
    _export_to_excel()
    return jsonify({"ok": True})


# ── STATS ─────────────────────────────────────────────────────────────────

@app.route("/api/stats/summary")
def stats_summary():
    logs  = read_json(LOGS_FILE, [])
    today = date.today()
    this_month = today.strftime("%Y-%m")
    this_week  = get_week_key(str(today))

    month_s = sum(1 for l in logs if l["log_date"][:7] == this_month)
    week_s  = sum(1 for l in logs if get_week_key(l["log_date"]) == this_week)
    vol     = sum((s.get("weight") or 0) * (s.get("reps") or 0) for l in logs for s in l.get("sets",[]))
    prs     = {}
    for l in logs:
        for s in l.get("sets",[]):
            ex_id = s.get("exercise_id"); w = s.get("weight") or 0
            if w > 0 and (ex_id not in prs or w > prs[ex_id]): prs[ex_id] = w
    return jsonify({"month_sessions": month_s, "week_sessions": week_s,
                    "total_volume": round(vol), "pr_count": len(prs), "total_sessions": len(logs)})

@app.route("/api/stats/personal_records")
def get_prs():
    logs = read_json(LOGS_FILE, [])
    exs  = read_json(EXERCISES_FILE, [])
    ex_map = {e["id"]: e for e in exs}
    prs = {}
    for l in logs:
        for s in l.get("sets",[]):
            ex_id = s.get("exercise_id"); w = s.get("weight") or 0
            if w > 0 and (ex_id not in prs or w > prs[ex_id]["max_weight"]):
                ex = ex_map.get(ex_id, {})
                prs[ex_id] = {"exercise_id": ex_id, "exercise_name": ex.get("name","?"),
                              "category": ex.get("category",""), "muscle_group": ex.get("muscle_group",""),
                              "max_weight": w, "reps_at_max": s.get("reps"), "last_performed": l["log_date"][:10]}
    return jsonify(sorted(prs.values(), key=lambda x: x["exercise_name"]))


# ── PLOTLY CHART ENDPOINTS ────────────────────────────────────────────────

@app.route("/api/charts/weekly_volume")
def chart_weekly_volume():
    logs = read_json(LOGS_FILE, [])
    week_vol  = defaultdict(float)
    week_sess = defaultdict(int)
    for l in logs:
        wk = get_week_key(l["log_date"])
        week_sess[wk] += 1
        for s in l.get("sets",[]):
            week_vol[wk] += (s.get("weight") or 0) * (s.get("reps") or 0)

    weeks = sorted(set(list(week_vol.keys()) + list(week_sess.keys())))[-16:]
    labels = [f"S{w.split('-W')[1]} '{w[:4][2:]}" for w in weeks]

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(go.Bar(
        x=labels, y=[round(week_vol[w]) for w in weeks], name="Volume (kg)",
        marker=dict(color="#e94560", opacity=0.85, line=dict(width=0)),
        hovertemplate="<b>%{x}</b><br>Volume : <b>%{y} kg</b><extra></extra>",
    ), secondary_y=False)
    fig.add_trace(go.Scatter(
        x=labels, y=[week_sess[w] for w in weeks], name="Séances",
        mode="lines+markers",
        line=dict(color="#00d4ff", width=2.5, dash="solid"),
        marker=dict(color="#00d4ff", size=7, line=dict(color="#111118", width=1.5)),
        hovertemplate="<b>%{x}</b><br>Séances : <b>%{y}</b><extra></extra>",
    ), secondary_y=True)

    fig.update_layout(
        title=dict(text="Volume & Séances Hebdomadaires", font=dict(color="#f0f0f8", size=14)),
        barmode="overlay",
        **{k: v for k, v in LAYOUT_BASE.items() if k not in ("colorway",)},
    )
    fig.update_yaxes(title_text="Volume (kg)", secondary_y=False, **AXIS_STYLE)
    fig.update_yaxes(title_text="Séances", secondary_y=True,
                     tickfont=dict(color="#00d4ff"), gridcolor="#1e1e2a", linecolor="#2a2a38")
    fig.update_xaxes(**AXIS_STYLE)
    return jsonify(json.loads(fig.to_json()))


@app.route("/api/charts/muscle_distribution")
def chart_muscle_dist():
    logs = read_json(LOGS_FILE, [])
    exs  = read_json(EXERCISES_FILE, [])
    ex_map = {e["id"]: e for e in exs}
    muscle_vol = defaultdict(float)
    for l in logs:
        for s in l.get("sets",[]):
            ex = ex_map.get(s.get("exercise_id"), {})
            muscle_vol[ex.get("muscle_group","Autre")] += (s.get("weight") or 0) * (s.get("reps") or 0)

    if not muscle_vol:
        return jsonify({})

    items = sorted(muscle_vol.items(), key=lambda x: x[1], reverse=True)[:8]
    fig = go.Figure(go.Pie(
        labels=[i[0] for i in items],
        values=[round(i[1]) for i in items],
        hole=0.58,
        textinfo="label+percent",
        textfont=dict(size=11, color="#d0d0e8"),
        marker=dict(
            colors=["#e94560","#00d4ff","#ffd740","#00e676","#c792ea","#ff6b6b","#82aaff","#ffab40"],
            line=dict(color="#111118", width=2),
        ),
        hovertemplate="<b>%{label}</b><br>%{value} kg<br>%{percent}<extra></extra>",
    ))
    fig.update_layout(
        title=dict(text="Répartition du Volume par Muscle", font=dict(color="#f0f0f8", size=14)),
        **LAYOUT_BASE,
    )
    return jsonify(json.loads(fig.to_json()))


@app.route("/api/charts/exercise_progress/<ex_id>")
def chart_exercise_progress(ex_id):
    logs = read_json(LOGS_FILE, [])
    exs  = read_json(EXERCISES_FILE, [])
    ex_map = {e["id"]: e for e in exs}
    ex = ex_map.get(ex_id)
    if not ex: abort(404)

    is_cali = ex.get("category") == "calisthenics"
    date_data = defaultdict(lambda: {"max_weight": 0, "max_reps": 0, "volume": 0})

    for l in logs:
        for s in l.get("sets",[]):
            if s.get("exercise_id") != ex_id: continue
            d = l["log_date"][:10]
            w = s.get("weight") or 0; r = s.get("reps") or 0
            date_data[d]["max_weight"] = max(date_data[d]["max_weight"], w)
            date_data[d]["max_reps"]   = max(date_data[d]["max_reps"], r)
            date_data[d]["volume"]    += w * r

    if not date_data: return jsonify({})

    dates = sorted(date_data.keys())
    fig = make_subplots(rows=2, cols=1, shared_xaxes=True,
                        subplot_titles=["Poids Max (kg)" if not is_cali else "Reps Max", "Volume Total (kg)"],
                        vertical_spacing=0.14, row_heights=[0.6, 0.4])

    y_top = [date_data[d]["max_weight"] for d in dates] if not is_cali else [date_data[d]["max_reps"] for d in dates]
    c = "#e94560" if not is_cali else "#00e676"
    rgb = {"#e94560": "233,69,96", "#00e676": "0,230,118"}[c]

    fig.add_trace(go.Scatter(
        x=dates, y=y_top,
        mode="lines+markers",
        name="Poids Max" if not is_cali else "Reps Max",
        line=dict(color=c, width=2.5),
        marker=dict(color=c, size=7, line=dict(color="#111118", width=1.5)),
        fill="tozeroy", fillcolor=f"rgba({rgb},0.1)",
        hovertemplate="%{x}<br><b>%{y:.1f}" + (" kg" if not is_cali else " reps") + "</b><extra></extra>",
    ), row=1, col=1)

    fig.add_trace(go.Bar(
        x=dates, y=[round(date_data[d]["volume"]) for d in dates],
        name="Volume",
        marker=dict(color="#00d4ff", opacity=0.75, line=dict(width=0)),
        hovertemplate="%{x}<br>Volume : <b>%{y} kg</b><extra></extra>",
    ), row=2, col=1)

    fig.update_layout(
        title=dict(text=f"Progression — {ex.get('name','')}", font=dict(color="#f0f0f8", size=14)),
        showlegend=True, **LAYOUT_BASE,
    )
    for ax in ["xaxis","xaxis2","yaxis","yaxis2"]:
        fig.update_layout(**{ax: AXIS_STYLE})
    return jsonify(json.loads(fig.to_json()))


@app.route("/api/charts/heatmap")
def chart_heatmap():
    logs = read_json(LOGS_FILE, [])
    exs  = read_json(EXERCISES_FILE, [])
    ex_map = {e["id"]: e for e in exs}

    data = defaultdict(lambda: defaultdict(float))
    for l in logs:
        wk = get_week_key(l["log_date"])
        for s in l.get("sets",[]):
            ex = ex_map.get(s.get("exercise_id"), {})
            mg = ex.get("muscle_group","Autre")
            data[wk][mg] += (s.get("weight") or 0) * (s.get("reps") or 0)

    if not data: return jsonify({})

    weeks   = sorted(data.keys())[-14:]
    muscles = sorted({mg for wk in weeks for mg in data[wk].keys()})
    z = [[round(data[wk].get(mg, 0)) for wk in weeks] for mg in muscles]
    xlabels = [f"S{w.split('-W')[1]}" for w in weeks]

    fig = go.Figure(go.Heatmap(
        z=z, x=xlabels, y=muscles,
        colorscale=[[0,"#0a0a18"],[0.15,"#1a1a3e"],[0.5,"#e94560"],[1,"#ffd740"]],
        hovertemplate="Sem. %{x} — %{y}<br><b>%{z} kg</b><extra></extra>",
        colorbar=dict(tickfont=dict(color="#8888aa", size=10),
                      title=dict(text="kg", font=dict(color="#8888aa"))),
    ))
    fig.update_layout(
        title=dict(text="Heatmap Volume — Muscle × Semaine", font=dict(color="#f0f0f8", size=14)),
        **LAYOUT_BASE,
    )
    return jsonify(json.loads(fig.to_json()))


@app.route("/api/charts/pr_bars")
def chart_pr_bars():
    logs = read_json(LOGS_FILE, [])
    exs  = read_json(EXERCISES_FILE, [])
    ex_map = {e["id"]: e for e in exs}
    prs = {}
    for l in logs:
        for s in l.get("sets",[]):
            ex_id = s.get("exercise_id"); w = s.get("weight") or 0
            if w > 0 and (ex_id not in prs or w > prs[ex_id]["weight"]):
                ex = ex_map.get(ex_id, {})
                prs[ex_id] = {"name": ex.get("name","?"), "weight": w, "reps": s.get("reps",0),
                               "muscle": ex.get("muscle_group","Autre"), "date": l["log_date"][:10]}
    if not prs: return jsonify({})

    items = sorted(prs.values(), key=lambda x: x["weight"], reverse=True)[:15]
    MCOL = {"Pectoraux":"#e94560","Dos":"#00d4ff","Jambes":"#00e676","Épaules":"#ffd740",
             "Biceps":"#c792ea","Triceps":"#ff6b6b","Fessiers":"#82aaff","Mollets":"#ffab40",
             "Abdominaux":"#80cbc4","Full Body":"#ff80ab"}

    fig = go.Figure(go.Bar(
        y=[i["name"] for i in items],
        x=[i["weight"] for i in items],
        orientation="h",
        marker=dict(color=[MCOL.get(i["muscle"],"#8888aa") for i in items], line=dict(width=0),
                    opacity=0.9),
        text=[f"  {i['weight']} kg" for i in items],
        textposition="outside", textfont=dict(color="#c0c0d0", size=11),
        customdata=[f"<b>{i['name']}</b><br>{i['weight']} kg × {i['reps']} reps<br>{i['date']}" for i in items],
        hovertemplate="%{customdata}<extra></extra>",
    ))
    fig.update_layout(
        title=dict(text="Top 15 Records Personnels", font=dict(color="#f0f0f8", size=14)),
        height=max(360, len(items) * 30 + 90),
        xaxis_title="Poids Max (kg)",
        **LAYOUT_BASE,
    )
    fig.update_xaxes(**AXIS_STYLE)
    fig.update_yaxes(**AXIS_STYLE)
    return jsonify(json.loads(fig.to_json()))


# ── EXCEL EXPORT ───────────────────────────────────────────────────────────

def _export_to_excel():
    logs = read_json(LOGS_FILE, [])
    exs  = read_json(EXERCISES_FILE, [])
    progs = read_json(PROGRAMS_FILE, [])
    ex_map   = {e["id"]: e for e in exs}
    prog_map = {p["id"]: p["name"] for p in progs}
    sess_map = {s["id"]: s["name"] for p in progs for s in p["sessions"]}

    wb = openpyxl.Workbook()
    BG  = PatternFill("solid", fgColor="0a0a0f")
    BG2 = PatternFill("solid", fgColor="16161f")
    BG3 = PatternFill("solid", fgColor="1e1e2a")
    RED  = PatternFill("solid", fgColor="e94560")
    GOLD = PatternFill("solid", fgColor="b8860b")
    TEAL = PatternFill("solid", fgColor="006064")

    thin = Border(left=Side(style="thin",color="2a2a38"), right=Side(style="thin",color="2a2a38"),
                  top=Side(style="thin",color="2a2a38"), bottom=Side(style="thin",color="2a2a38"))
    ctr = Alignment(horizontal="center", vertical="center")

    def cell_style(ws, r, c, fill, font, border=thin, align=ctr):
        cell = ws.cell(row=r, column=c)
        cell.fill = fill; cell.font = font; cell.border = border; cell.alignment = align
        return cell

    # ─ Sheet 1 ─
    ws = wb.active; ws.title = "📋 Historique"; ws.sheet_view.showGridLines = False
    ws['A1'] = "🏋️  FITTRACKER — HISTORIQUE"; ws['A1'].font = Font(name="Calibri",bold=True,color="FFFFFF",size=14)
    ws['A1'].fill = RED; ws['A1'].alignment = ctr; ws.merge_cells("A1:I1"); ws.row_dimensions[1].height = 30

    hdrs = ["Date","Programme","Séance","Exercice","Catégorie","Set","Reps","Poids (kg)","Volume (kg)"]
    for i,h in enumerate(hdrs,1):
        cell_style(ws,2,i,BG,Font(name="Calibri",bold=True,color="e94560",size=11)).value = h
    ws.row_dimensions[2].height = 22

    row_n = 3
    for log in sorted(logs, key=lambda l: l.get("log_date",""), reverse=True):
        for s in log.get("sets",[]):
            ex = ex_map.get(s.get("exercise_id"),{})
            w = s.get("weight") or 0; r2 = s.get("reps") or 0
            vals = [log["log_date"][:10], prog_map.get(log.get("program_id"),"—"),
                    sess_map.get(log.get("session_id"),"—"), ex.get("name","?"),
                    ex.get("category",""), s.get("set_number",""),
                    s.get("figure") or r2, w or "", round(w*r2,1) if w and r2 else ""]
            fill = BG3 if row_n % 2 == 0 else BG2
            for j,v in enumerate(vals,1):
                cell_style(ws,row_n,j,fill,Font(name="Calibri",color="e0e0f0",size=10)).value = v
            row_n += 1

    for col,w in zip("ABCDEFGHI",[14,24,24,28,14,6,8,12,12]):
        ws.column_dimensions[col].width = w

    # ─ Sheet 2 Records ─
    ws2 = wb.create_sheet("🏆 Records"); ws2.sheet_view.showGridLines = False
    ws2['A1'] = "🏆  PERSONAL RECORDS"; ws2['A1'].font = Font(name="Calibri",bold=True,color="FFFFFF",size=14)
    ws2['A1'].fill = GOLD; ws2['A1'].alignment = ctr; ws2.merge_cells("A1:F1"); ws2.row_dimensions[1].height = 30
    for i,h in enumerate(["Exercice","Catégorie","Muscle","Poids Max (kg)","Reps","Dernière perf"],1):
        cell_style(ws2,2,i,BG,Font(name="Calibri",bold=True,color="ffd740",size=11)).value = h

    prs = {}
    for l in logs:
        for s in l.get("sets",[]):
            ex_id = s.get("exercise_id"); w = s.get("weight") or 0
            if w > 0 and (ex_id not in prs or w > prs[ex_id]["max_weight"]):
                ex = ex_map.get(ex_id,{})
                prs[ex_id] = {"name":ex.get("name","?"),"cat":ex.get("category",""),"mg":ex.get("muscle_group",""),
                               "max_weight":w,"reps":s.get("reps"),"date":l["log_date"][:10]}
    for i,pr in enumerate(sorted(prs.values(),key=lambda x:x["name"]),3):
        fill = BG3 if i%2==0 else BG2
        for j,v in enumerate([pr["name"],pr["cat"],pr["mg"],pr["max_weight"],pr["reps"],pr["date"]],1):
            cell_style(ws2,i,j,fill,Font(name="Calibri",color="e0e0f0",size=10)).value = v
    for col,w in zip("ABCDEF",[28,14,20,16,8,14]):
        ws2.column_dimensions[col].width = w

    # ─ Sheet 3 Semaines ─
    ws3 = wb.create_sheet("📊 Semaines"); ws3.sheet_view.showGridLines = False
    ws3['A1'] = "📊  RÉSUMÉ HEBDOMADAIRE"; ws3['A1'].font = Font(name="Calibri",bold=True,color="FFFFFF",size=14)
    ws3['A1'].fill = TEAL; ws3['A1'].alignment = ctr; ws3.merge_cells("A1:D1"); ws3.row_dimensions[1].height = 30
    for i,h in enumerate(["Semaine","Séances","Volume (kg)","Jours"],1):
        cell_style(ws3,2,i,BG,Font(name="Calibri",bold=True,color="00d4ff",size=11)).value = h

    week_data = defaultdict(lambda: {"sessions":0,"volume":0.0,"days":set()})
    for l in logs:
        wk = get_week_key(l["log_date"])
        week_data[wk]["sessions"] += 1
        week_data[wk]["days"].add(l["log_date"][:10])
        for s in l.get("sets",[]):
            week_data[wk]["volume"] += (s.get("weight") or 0)*(s.get("reps") or 0)
    for i,wk in enumerate(sorted(week_data.keys(),reverse=True),3):
        d = week_data[wk]; fill = BG3 if i%2==0 else BG2
        for j,v in enumerate([wk,d["sessions"],round(d["volume"]),len(d["days"])],1):
            cell_style(ws3,i,j,fill,Font(name="Calibri",color="e0e0f0",size=10)).value = v
    for col,w in zip("ABCD",[14,10,16,10]):
        ws3.column_dimensions[col].width = w

    wb.save(EXCEL_FILE)


@app.route("/api/export/excel")
def export_excel():
    _export_to_excel()
    return send_from_directory(str(DATA_DIR), "fittracker_export.xlsx", as_attachment=True)


# ── ENTRY POINT ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    init_data()
    print("\n✅  FitTracker démarré → http://localhost:5000\n")
    app.run(debug=True, port=5000, use_reloader=False)